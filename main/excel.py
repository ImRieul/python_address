import openpyxl
import platform

# openpyxl는 직관성이 떨어져 읽기 쉬운 방식으로 만들어봤다.
import setting
from error.error_excel import *


class Excel:
    def __init__(self, name, read_only=False):
        self.__name = name
        self.__read_only = read_only
        self.workbook = None

        self.__read_excel()

    def __read_excel(self):
        file_path = f'{setting.ROOT_PROJECT}/{self.__name}'\
                    if platform.system() != 'Windows'\
                    else f'{setting.ROOT_PROJECT}\\{self.__name}'

        if not isinstance(self.__name, str):
            raise ExcelFileNameTypeError
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError as e:
            if self.__read_only:
                raise e
            openpyxl.Workbook().save(file_path)
            self.workbook = openpyxl.load_workbook(file_path)

    def get_file_name(self):
        return self.__name

    def get_sheet_name(self, name=None):
        if name is not None and not isinstance(name, str):
            raise ExcelSheetNameTypeErrors
        elif name is None:
            return self.workbook.active.title
        else:
            return self.workbook[name].title

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def set_sheet_name(self, name=None):
        # if
        self.workbook.active.title = name
        self.workbook.save(self.__name)

    def get_active_sheet(self, name=None):
        return self.workbook.active\
                if name is not None else self.workbook[name]\
                if type(name, str) else None


if __name__ == '__main__':
    def function():
        pass
    # os.chdir('../')
    excel = Excel('sample.xlsx')
    excel.get_sheet_name(function)

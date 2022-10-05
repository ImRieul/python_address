from __future__ import annotations

import openpyxl
import platform

import setting


# openpyxl는 직관성이 떨어져 읽기 쉬운 방식으로 만들어봤다.
# TODO
#   1. getter, setter를 python의 방식대로 수정.

from error.error_excel import *


class Excel:
    def __init__(self, name, read_only=False):
        self.__name = name
        self.__read_only = read_only
        self.workbook = None
        # self.sheet_name = None

        self.__read_excel()

    # 모든 작업을 마친 후 저장
    def __exit__(self, exc_type, exc_val, exc_tb):
        self.workbook.save(self.__name)

    def __read_excel(self):
        file_path = f'{setting.ROOT_PROJECT}/{self.__name}' \
            if platform.system() != 'Windows' \
            else f'{setting.ROOT_PROJECT}\\{self.__name}'

        # if not isinstance(self.__name, str):
        #     raise ExcelFileNameTypeError
        try:
            self.workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError as e:
            if self.__read_only:
                raise e
            openpyxl.Workbook().save(file_path)
            self.workbook = openpyxl.load_workbook(file_path)

    def get_file_name(self):
        return self.__name

    def get_sheet_name(self, name: str | None = None) -> str:
        # if name is not None and not isinstance(name, str):
        #     raise ExcelSheetNameTypeError
        # elif name is None:
        if name is None:
            return self.workbook.active.title
        else:
            return self.workbook[name].title

    def get_sheet_names(self):
        return self.workbook.sheetnames

    def set_sheet_name(self, name=None):
        # if
        self.workbook.active.title = name
        self.workbook.save(self.__name)

    def get_sheet_active(self, name: str | None = None):
        return self.workbook.active \
            if name is not None else self.workbook[name]

    # 앞으로 개발할 기능들
    def __copy_sheet(self, name):
        rename = f'copy_{name}'  # 이후 copy가 중복된다면 숫자를 붙일 수 있도록.
        self.workbook.copy_worksheet(self.workbook[name])
        self.workbook[f'{name} Copy'].title = f'copy_{name}'


if __name__ == '__main__':
    def function():
        pass
    # os.chdir('../')
    excel = Excel('sample.xlsx')

    # 열을 알파벳으로 하면 반복문으로 돌리기가 어려워진다.
    # pandas같은 라이브러리를 이용해 불러오는 순간 데이터화하는 건 어떨까?
    column_2 = excel.workbook['data'][2]
    row_a = excel.workbook['data']['A']
